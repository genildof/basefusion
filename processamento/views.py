from django.shortcuts import render, get_object_or_404, redirect
import os
import pandas as pd
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.utils import timezone
from .models import ArquivoProcessado, RegistroRevisao, BaseConsolidada, PerfilUsuario, LogAlteracaoPendenciaMacro
from django.views.decorators.csrf import csrf_exempt
import json
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView
from django.db.models import Q, Count
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator
import tempfile
from django.core.files.storage import FileSystemStorage
from datetime import datetime
from django.db import transaction
import numpy as np
import warnings
from django.contrib import messages
from django.views import View
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db.models.functions import TruncDate
from .modules.custom_filters import get_filtro
from .modules.database_updates import atualizar_pedidos_report_b2b, atualizar_rede_externa
from .modules.business_rules import processar_regras_negocio
import openpyxl
import time
from django.db import models
from django.urls import reverse_lazy
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django import forms
from django.contrib.auth import logout
from .modules.constants import PENDENCIAS_MACRO_OPCOES
from pathlib import Path

# Definir BASE_DIR
BASE_DIR = Path(__file__).resolve().parent.parent

# Verificações de perfil de usuário
def is_supervisor(user):
    try:
        return user.perfil.is_supervisor
    except:
        return False

# Formulários para usuários
class UsuarioForm(UserCreationForm):
    email = forms.EmailField(required=True)
    first_name = forms.CharField(max_length=30, required=True, label="Nome")
    last_name = forms.CharField(max_length=30, required=True, label="Sobrenome")
    
    class Meta:
        model = User
        fields = ('username', 'email', 'first_name', 'last_name', 'password1', 'password2')
    
    def save(self, commit=True):
        user = super().save(commit=False)
        user.email = self.cleaned_data['email']
        user.first_name = self.cleaned_data['first_name']
        user.last_name = self.cleaned_data['last_name']
        
        if commit:
            user.save()
        return user

class PerfilUsuarioForm(forms.ModelForm):
    class Meta:
        model = PerfilUsuario
        fields = ['perfil']
        widgets = {
            'perfil': forms.Select(attrs={'class': 'form-select'})
        }

# Create your views here.

@login_required
def home(request):
    return render(request, 'processamento/home.html')

# Views de gerenciamento de usuários
@login_required
@user_passes_test(is_supervisor)
def lista_usuarios(request):
    usuarios = User.objects.all().order_by('username')
    
    # Verifica se cada usuário tem um perfil, caso não tenha, cria um perfil padrão
    for usuario in usuarios:
        try:
            perfil = usuario.perfil
        except:
            PerfilUsuario.objects.create(usuario=usuario, perfil='OPERADOR')
    
    # Recarrega os usuários com seus perfis
    usuarios = User.objects.all().select_related('perfil').order_by('username')
    
    # Conta o número de supervisores
    total_supervisores = PerfilUsuario.objects.filter(perfil='SUPERVISOR').count()
    
    context = {
        'usuarios': usuarios,
        'total_supervisores': total_supervisores
    }
    return render(request, 'processamento/usuarios/lista_usuarios.html', context)

@login_required
@user_passes_test(is_supervisor)
def criar_usuario(request):
    if request.method == 'POST':
        form_usuario = UsuarioForm(request.POST)
        form_perfil = PerfilUsuarioForm(request.POST)
        
        if form_usuario.is_valid() and form_perfil.is_valid():
            with transaction.atomic():
                usuario = form_usuario.save()
                perfil = form_perfil.save(commit=False)
                perfil.usuario = usuario
                perfil.save()
                
                messages.success(request, 'Usuário criado com sucesso!')
                return redirect('lista_usuarios')
    else:
        form_usuario = UsuarioForm()
        form_perfil = PerfilUsuarioForm()
    
    context = {
        'form_usuario': form_usuario,
        'form_perfil': form_perfil,
        'titulo': 'Adicionar Usuário'
    }
    return render(request, 'processamento/usuarios/form_usuario.html', context)

@login_required
@user_passes_test(is_supervisor)
def editar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    
    try:
        perfil = usuario.perfil
    except:
        perfil = PerfilUsuario.objects.create(usuario=usuario, perfil='OPERADOR')
    
    if request.method == 'POST':
        form_perfil = PerfilUsuarioForm(request.POST, instance=perfil)
        
        if form_perfil.is_valid():
            form_perfil.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('lista_usuarios')
    else:
        form_perfil = PerfilUsuarioForm(instance=perfil)
    
    context = {
        'usuario': usuario,
        'form_perfil': form_perfil,
        'titulo': 'Editar Usuário'
    }
    return render(request, 'processamento/usuarios/editar_usuario.html', context)

@login_required
@user_passes_test(is_supervisor)
def excluir_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, 'Usuário excluído com sucesso!')
        return redirect('lista_usuarios')
    
    context = {
        'usuario': usuario
    }
    return render(request, 'processamento/usuarios/excluir_usuario.html', context)

# Middleware de verificação de perfil
class PerfilMiddleware:
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        # Código executado antes da visualização ser chamada
        if request.user.is_authenticated:
            try:
                # Verifica se o usuário tem um perfil
                perfil = request.user.perfil
                
                # Se o usuário for operador, só pode acessar a página de revisão
                if perfil.is_operador:
                    allowed_urls = [
                        '/revisao/', 
                        '/revisao/data/', 
                        '/revisao/segmentos/', 
                        '/revisao/registro/',
                        '/accounts/logout/'
                    ]
                    
                    # Verifica se a URL atual está na lista de permitidas
                    path = request.path
                    allowed = False
                    
                    for url in allowed_urls:
                        if path.startswith(url):
                            allowed = True
                            break
                    
                    # Se não estiver na lista de permitidas e não for a página de login, redireciona
                    if not allowed and not path.startswith('/accounts/login/') and not path == '/':
                        return redirect('revisao')
            except:
                # Se o usuário não tem perfil, cria um perfil padrão
                PerfilUsuario.objects.create(usuario=request.user, perfil='OPERADOR')
        
        # Processa a requisição
        response = self.get_response(request)
        
        return response

@login_required
def upload_arquivos(request):
    if request.method == 'POST':
        if 'arquivo' not in request.FILES and 'processar_regras' not in request.POST:
            return JsonResponse({
                'success': False,
                'message': 'Nenhum arquivo foi enviado.'
            }, status=400)
            
        try:
            if 'processar_regras' in request.POST:
                try:
                    start_time = time.time()
                    
                    # Registra o início do processamento
                    log = ArquivoProcessado(
                        nome="Processamento de Regras de Negócio",
                        tipo_arquivo="REGRAS_NEGOCIO",
                        status="PROCESSANDO",
                        usuario=request.user
                    )
                    log.save()
                    
                    # Obtém os registros da BaseConsolidada
                    registros = BaseConsolidada.objects.all()
                    
                    # Converte para DataFrame
                    df = pd.DataFrame(list(registros.values()))
                    
                    # Processa as regras de negócio
                    resultado = processar_regras_negocio(df)
                    
                    if resultado['success']:
                        df_atualizado = resultado['df_atualizado']
                        total_atualizados = resultado['resultados']['total_atualizados']
                        
                        # Atualizar o banco de dados
                        print("Atualizando banco de dados...")
                        atualizados = 0
                        erros = 0
                        
                        # Processar cada registro individualmente
                        for pedido, row in df_atualizado.iterrows():
                            try:
                                # Usar transação atômica para cada registro
                                with transaction.atomic():
                                    registro = BaseConsolidada.objects.get(pedido=pedido)
                                    
                                    # Atualiza os campos modificados
                                    if 'tipo_entrega' in row and not pd.isna(row['tipo_entrega']):
                                        registro.tipo_entrega = row['tipo_entrega']
                                    if 'data_sae' in row and not pd.isna(row['data_sae']):
                                        registro.data_sae = row['data_sae']
                                    if 'esteira' in row and not pd.isna(row['esteira']):
                                        registro.esteira = row['esteira']
                                    if 'segmento_v3' in row and not pd.isna(row['segmento_v3']):
                                        registro.segmento_v3 = row['segmento_v3']
                                    if 'status_rede' in row and not pd.isna(row['status_rede']):
                                        registro.status_rede = row['status_rede']
                                    if 'agrupado' in row and not pd.isna(row['agrupado']):
                                        registro.agrupado = row['agrupado']
                                    if 'pendencia_macro' in row and not pd.isna(row['pendencia_macro']):
                                        registro.pendencia_macro = row['pendencia_macro']
                                    if 'estimativa' in row and not pd.isna(row['estimativa']):
                                        registro.estimativa = row['estimativa']
                                    
                                    registro.save()
                                    atualizados += 1
                                    
                                    if atualizados % 1000 == 0:
                                        print(f"Atualizados {atualizados} registros...")
                                        
                            except BaseConsolidada.DoesNotExist:
                                print(f"Registro não encontrado para o pedido: {pedido}")
                                erros += 1
                                continue
                            except Exception as e:
                                print(f"Erro ao atualizar registro {pedido}: {str(e)}")
                                erros += 1
                                continue
                        
                        # Calcula o tempo de processamento
                        end_time = time.time()
                        tempo_processamento = end_time - start_time
                        
                        # Atualiza o log com o resultado
                        log.status = "CONCLUIDO"
                        log.mensagem = f"Processamento de regras de negócio concluído com sucesso. Total de registros atualizados: {atualizados}, Erros: {erros}"
                        log.registros_afetados = atualizados
                        log.tempo_processamento = tempo_processamento
                        log.save()
                        
                        print(f"Processamento concluído. Total de registros atualizados: {atualizados}, Erros: {erros}")
                        messages.success(request, f"Regras de negócio processadas com sucesso. Total de registros atualizados: {atualizados}, Erros: {erros}")
                    else:
                        # Em caso de erro, registra no log
                        log.status = "ERRO"
                        log.mensagem = f"Erro ao processar regras de negócio: {resultado['error']}"
                        log.save()
                        
                        print(f"Erro durante o processamento: {resultado['error']}")
                        messages.error(request, f"Erro durante o processamento das regras de negócio: {resultado['error']}")
                    
                    return redirect('upload_arquivos')
                except Exception as e:
                    # Em caso de erro não tratado, registra no log
                    if 'log' in locals():
                        log.status = "ERRO"
                        log.mensagem = f"Erro ao processar regras de negócio: {str(e)}"
                        log.save()
                    else:
                        ArquivoProcessado.objects.create(
                            nome="Processamento de Regras de Negócio",
                            tipo_arquivo="REGRAS_NEGOCIO",
                            status="ERRO",
                            mensagem=f"Erro ao processar regras de negócio: {str(e)}",
                            usuario=request.user
                        )
                    
                    messages.error(request, f'Erro ao processar regras de negócio: {str(e)}')
                    return redirect('upload_arquivos')
            
            # Processamento normal de arquivo
            arquivo = request.FILES['arquivo']
            
            # Salvar o arquivo temporariamente
            temp_file = tempfile.NamedTemporaryFile(delete=False)
            temp_file.write(arquivo.read())
            temp_file.close()
            
            # Verificar o tipo de arquivo
            xls = pd.ExcelFile(temp_file.name)
            print(f"Abas disponíveis no arquivo: {xls.sheet_names}")
            
            # Verificar se é base de Rede Externa
            if 'ANALITICO' in xls.sheet_names:
                df = pd.read_excel(temp_file.name, sheet_name='ANALITICO')
                print(f"Colunas encontradas na aba ANALITICO: {df.columns.tolist()}")
                
                colunas_rede = ['PEDIDO', 'CONTRATADA', 'Prazo Rede']
                colunas_arquivo = [col.upper().strip() for col in df.columns]
                colunas_rede_norm = [col.upper().strip() for col in colunas_rede]
                
                if all(col in colunas_arquivo for col in colunas_rede_norm):
                    try:
                        start_time = time.time()
                        
                        # Registra o início do processamento
                        log = ArquivoProcessado(
                            nome=arquivo.name,
                            tipo_arquivo="REDE_EXTERNA",
                            status="PROCESSANDO",
                            usuario=request.user
                        )
                        log.save()
                        
                        stats = atualizar_rede_externa(temp_file.name)
                        
                        # Calcula o tempo de processamento
                        end_time = time.time()
                        tempo_processamento = end_time - start_time
                        
                        # Atualiza o log com o resultado
                        log.status = "CONCLUIDO"
                        log.mensagem = f"Arquivo de Rede Externa processado com sucesso. Total na base de Rede: {stats['total_rede']}, Total na base do sistema: {stats['total_base']}, Pedidos atualizados: {stats['atualizados']}"
                        log.registros_afetados = stats['atualizados']
                        log.tempo_processamento = tempo_processamento
                        log.save()
                        
                        messages.success(request, f'''
                            Arquivo de Rede Externa processado com sucesso!
                            Total na base de Rede: {stats['total_rede']}
                            Total na base do sistema: {stats['total_base']}
                            Pedidos atualizados: {stats['atualizados']}
                        ''')
                    except Exception as e:
                        if 'log' in locals():
                            log.status = "ERRO"
                            log.mensagem = f"Erro ao processar base de Rede Externa: {str(e)}"
                            log.save()
                        else:
                            ArquivoProcessado.objects.create(
                                nome=arquivo.name,
                                tipo_arquivo="REDE_EXTERNA",
                                status="ERRO",
                                mensagem=f"Erro ao processar base de Rede Externa: {str(e)}",
                                usuario=request.user
                            )
                        
                        raise Exception(f'Erro ao processar base de Rede Externa: {str(e)}')
                else:
                    ArquivoProcessado.objects.create(
                        nome=arquivo.name,
                        tipo_arquivo="REDE_EXTERNA",
                        status="ERRO",
                        mensagem=f"Colunas necessárias não encontradas na aba ANALITICO. Colunas encontradas: {df.columns.tolist()}, Colunas esperadas: {colunas_rede}",
                        usuario=request.user
                    )
                    
                    raise Exception(f'''
                        Colunas necessárias não encontradas na aba ANALITICO.
                        Colunas encontradas: {df.columns.tolist()}
                        Colunas esperadas: {colunas_rede}
                    ''')
            # Verificar se é report B2B
            elif 'Report' in xls.sheet_names:
                df = pd.read_excel(temp_file.name, sheet_name='Report')
                print(f"Colunas encontradas na aba Report: {df.columns.tolist()}")
                
                colunas_b2b = ['Pedido', 'Cliente', 'Cidade', 'Esteira']
                colunas_arquivo = [col.upper().strip() for col in df.columns]
                colunas_b2b_norm = [col.upper().strip() for col in colunas_b2b]
                
                if all(col in colunas_arquivo for col in colunas_b2b_norm):
                    try:
                        start_time = time.time()
                        
                        # Registra o início do processamento
                        log = ArquivoProcessado(
                            nome=arquivo.name,
                            tipo_arquivo="REPORT_B2B",
                            status="PROCESSANDO",
                            usuario=request.user
                        )
                        log.save()
                        
                        stats = atualizar_pedidos_report_b2b(temp_file.name)
                        
                        # Calcula o tempo de processamento
                        end_time = time.time()
                        tempo_processamento = end_time - start_time
                        
                        # Adiciona informações sobre colunas na mensagem
                        info_colunas = ""
                        if stats.get('colunas_faltantes'):
                            info_colunas = f"\nAtenção: {len(stats['colunas_faltantes'])} colunas esperadas não foram encontradas no arquivo."
                        
                        # Atualiza o log com o resultado
                        log.status = "CONCLUIDO"
                        log.mensagem = f"Arquivo B2B processado com sucesso. Total bruto: {stats['total_registros_brutos']}, Ignorados: {stats['total_ignorados']}, Total válidos: {stats['total_report']}, Total na base: {stats['total_base']}, Pedidos criados: {stats['criados']}, Pedidos atualizados: {stats['atualizados']}, Pedidos removidos: {stats['removidos']}, Registros inválidos removidos: {stats['registros_invalidos_removidos']}. Colunas importadas: {stats['colunas_importadas']}/30.{info_colunas}"
                        log.registros_afetados = stats['criados'] + stats['atualizados'] + stats['removidos'] + stats['registros_invalidos_removidos']
                        log.tempo_processamento = tempo_processamento
                        log.save()
                        
                        messages.success(request, f'''
                            Arquivo B2B processado com sucesso!
                            Total registros brutos no arquivo: {stats['total_registros_brutos']}
                            Linhas ignoradas (sem Esteira ou inválidas): {stats['total_ignorados']}
                            Total registros válidos processados: {stats['total_report']}
                            Total na base: {stats['total_base']}
                            Pedidos criados: {stats['criados']}
                            Pedidos atualizados: {stats['atualizados']}
                            Pedidos removidos: {stats['removidos']}
                            Registros inválidos na base removidos: {stats['registros_invalidos_removidos']}
                            Colunas importadas: {stats['colunas_importadas']}/30
                            {info_colunas}
                        ''')
                    except Exception as e:
                        if 'log' in locals():
                            log.status = "ERRO"
                            log.mensagem = f"Erro ao processar report B2B: {str(e)}"
                            log.save()
                        else:
                            ArquivoProcessado.objects.create(
                                nome=arquivo.name,
                                tipo_arquivo="REPORT_B2B",
                                status="ERRO",
                                mensagem=f"Erro ao processar report B2B: {str(e)}",
                                usuario=request.user
                            )
                        
                        raise Exception(f'Erro ao processar report B2B: {str(e)}')
                else:
                    ArquivoProcessado.objects.create(
                        nome=arquivo.name,
                        tipo_arquivo="REPORT_B2B",
                        status="ERRO",
                        mensagem=f"Colunas necessárias não encontradas na aba Report. Colunas encontradas: {df.columns.tolist()}, Colunas esperadas: {colunas_b2b}",
                        usuario=request.user
                    )
                    
                    raise Exception(f'''
                        Colunas necessárias não encontradas na aba Report.
                        Colunas encontradas: {df.columns.tolist()}
                        Colunas esperadas: {colunas_b2b}
                    ''')
            # Verificar se é aba Export (tentativa de tratar esse tipo de arquivo)
            elif 'Export' in xls.sheet_names:
                df = pd.read_excel(temp_file.name, sheet_name='Export')
                print(f"Colunas encontradas na aba Export: {df.columns.tolist()}")
                
                # Verificar se as colunas necessárias para um report B2B estão presentes
                colunas_b2b = ['Pedido', 'Cliente', 'Cidade', 'Esteira']
                colunas_arquivo = [col for col in df.columns]
                
                if all(col in colunas_arquivo for col in colunas_b2b):
                    try:
                        start_time = time.time()
                        
                        # Registra o início do processamento
                        log = ArquivoProcessado(
                            nome=arquivo.name,
                            tipo_arquivo="REPORT_B2B",
                            status="PROCESSANDO",
                            usuario=request.user
                        )
                        log.save()
                        
                        # Salvar o DataFrame com a aba Export para um arquivo temporário de Excel com aba Report
                        temp_b2b_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                        temp_b2b_file.close()
                        
                        # Criar um novo Excel com aba Report
                        with pd.ExcelWriter(temp_b2b_file.name, engine='openpyxl') as writer:
                            df.to_excel(writer, sheet_name='Report', index=False)
                        
                        # Processar como se fosse um report B2B
                        stats = atualizar_pedidos_report_b2b(temp_b2b_file.name)
                        
                        # Limpar o arquivo temporário
                        os.unlink(temp_b2b_file.name)
                        
                        # Calcula o tempo de processamento
                        end_time = time.time()
                        tempo_processamento = end_time - start_time
                        
                        # Adiciona informações sobre colunas na mensagem
                        info_colunas = ""
                        if stats.get('colunas_faltantes'):
                            info_colunas = f"\nAtenção: {len(stats['colunas_faltantes'])} colunas esperadas não foram encontradas no arquivo."
                        
                        # Atualiza o log com o resultado
                        log.status = "CONCLUIDO"
                        log.mensagem = f"Arquivo com aba 'Export' processado como B2B. Total bruto: {stats['total_registros_brutos']}, Ignorados: {stats['total_ignorados']}, Total válidos: {stats['total_report']}, Total na base: {stats['total_base']}, Pedidos criados: {stats['criados']}, Pedidos atualizados: {stats['atualizados']}, Pedidos removidos: {stats['removidos']}, Registros inválidos removidos: {stats['registros_invalidos_removidos']}. Colunas importadas: {stats['colunas_importadas']}/30.{info_colunas}"
                        log.registros_afetados = stats['criados'] + stats['atualizados'] + stats['removidos'] + stats['registros_invalidos_removidos']
                        log.tempo_processamento = tempo_processamento
                        log.save()
                        
                        messages.success(request, f'''
                            Arquivo com aba 'Export' processado como B2B!
                            Total registros brutos no arquivo: {stats['total_registros_brutos']}
                            Linhas ignoradas (sem Esteira ou inválidas): {stats['total_ignorados']}
                            Total registros válidos processados: {stats['total_report']}
                            Total na base: {stats['total_base']}
                            Pedidos criados: {stats['criados']}
                            Pedidos atualizados: {stats['atualizados']}
                            Pedidos removidos: {stats['removidos']}
                            Registros inválidos na base removidos: {stats['registros_invalidos_removidos']}
                            Colunas importadas: {stats['colunas_importadas']}/30
                            {info_colunas}
                        ''')
                        
                        # Remover arquivo temporário original
                        os.unlink(temp_file.name)
                        return JsonResponse({'success': True})
                        
                    except Exception as e:
                        # Limpar o arquivo temporário se ele existir
                        if 'temp_b2b_file' in locals():
                            try:
                                os.unlink(temp_b2b_file.name)
                            except:
                                pass
                        
                        if 'log' in locals():
                            log.status = "ERRO"
                            log.mensagem = f"Erro ao processar arquivo Export como B2B: {str(e)}"
                            log.save()
                        else:
                            ArquivoProcessado.objects.create(
                                nome=arquivo.name,
                                tipo_arquivo="REPORT_B2B",
                                status="ERRO",
                                mensagem=f"Erro ao processar arquivo Export como B2B: {str(e)}",
                                usuario=request.user
                            )
                        
                        # Remover arquivo temporário original
                        os.unlink(temp_file.name)
                        return JsonResponse({
                            'success': False,
                            'message': f"Erro ao processar arquivo Export como B2B: {str(e)}"
                        }, status=400)
                else:
                    # As colunas necessárias não estão presentes
                    ArquivoProcessado.objects.create(
                        nome=arquivo.name,
                        tipo_arquivo="REPORT_B2B",
                        status="ERRO",
                        mensagem=f"Arquivo com aba 'Export' não contém todas as colunas necessárias. Colunas encontradas: {colunas_arquivo}, Colunas esperadas: {colunas_b2b}",
                        usuario=request.user
                    )
                    
                    # Remover arquivo temporário
                    os.unlink(temp_file.name)
                    return JsonResponse({
                        'success': False,
                        'message': f'''
                            Arquivo com aba 'Export' não contém todas as colunas necessárias.
                            Colunas necessárias: {colunas_b2b}
                            Por favor, use um arquivo que contenha estas colunas.
                        '''
                    }, status=400)
            else:
                abas_disponíveis = ", ".join(xls.sheet_names)
                
                ArquivoProcessado.objects.create(
                    nome=arquivo.name,
                    tipo_arquivo="DESCONHECIDO",
                    status="ERRO",
                    mensagem=f"Formato de arquivo desconhecido. Abas encontradas: {abas_disponíveis}, Abas esperadas: ANALITICO (para Rede Externa) ou Report (para B2B).",
                    usuario=request.user
                )
                
                raise Exception(f'''
                    Nenhuma aba válida encontrada no arquivo.
                    Abas disponíveis: {xls.sheet_names}
                    Abas esperadas: ANALITICO (para Rede Externa) ou Report (para B2B)
                ''')
            
            # Remover arquivo temporário
            os.unlink(temp_file.name)
            return JsonResponse({'success': True})
            
        except Exception as e:
            # Remover arquivo temporário em caso de erro
            if 'temp_file' in locals():
                os.unlink(temp_file.name)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return render(request, 'processamento/upload.html')

@login_required
def lista_arquivos(request):
    # Obtém os filtros da requisição
    tipo_arquivo = request.GET.get('tipo_arquivo', '')
    status = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    # Filtra os arquivos processados
    arquivos = ArquivoProcessado.objects.all()
    
    if tipo_arquivo:
        arquivos = arquivos.filter(tipo_arquivo=tipo_arquivo)
    
    if status:
        arquivos = arquivos.filter(status=status)
    
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
        arquivos = arquivos.filter(data_processamento__gte=data_inicio)
    
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d')
        data_fim = data_fim.replace(hour=23, minute=59, second=59)
        arquivos = arquivos.filter(data_processamento__lte=data_fim)
    
    # Ordena por data de processamento (mais recente primeiro)
    arquivos = arquivos.order_by('-data_processamento')
    
    # Estatísticas
    total_arquivos = arquivos.count()
    total_registros_afetados = arquivos.filter(registros_afetados__isnull=False).aggregate(
        total=models.Sum('registros_afetados')
    )['total'] or 0
    
    # Resumo por tipo
    resumo_tipo = arquivos.values('tipo_arquivo').annotate(
        total=models.Count('id'),
        sucesso=models.Count('id', filter=models.Q(status='CONCLUIDO')),
        erro=models.Count('id', filter=models.Q(status='ERRO'))
    )
    
    contexto = {
        'arquivos': arquivos,
        'tipos_arquivo': ArquivoProcessado.TIPO_CHOICES,
        'status_opcoes': ArquivoProcessado.STATUS_CHOICES,
        'tipo_arquivo_filtro': tipo_arquivo,
        'status_filtro': status,
        'data_inicio_filtro': data_inicio,
        'data_fim_filtro': data_fim,
        'total_arquivos': total_arquivos,
        'total_registros_afetados': total_registros_afetados,
        'resumo_tipo': resumo_tipo
    }
    
    return render(request, 'processamento/lista_arquivos.html', contexto)

@login_required
def revisao(request):
    return render(request, 'processamento/revisao.html')

@login_required
def revisao_data(request):
    # Obter parâmetros do DataTable
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search = request.GET.get('search[value]', '')
    segmento = request.GET.get('segmento', '')
    tipo = request.GET.get('tipo', 'revisar')
    
    # Obter parâmetros de ordenação
    order_column = request.GET.get('order[0][column]', '0')
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    # Mapear colunas para campos do modelo
    column_map = {
        '0': 'pedido',
        '1': 'dias_carteira_atual',
        '2': 'data_tecnica',
        '3': 'carteira',
        '4': 'pendencia_macro',
        '5': 'estimativa',
        '6': 'segmento_v3'
    }
    
    # Obter filtros do módulo
    filtros = get_filtro(tipo)
    
    # Construir a query base
    query = Q()
    
    # Aplicar filtros baseado no tipo selecionado
    if 'pendencia_macro' in filtros:
        query &= Q(pendencia_macro__in=filtros['pendencia_macro'])
    if 'classificacao_resumo_atual' in filtros:
        query &= Q(classificacao_resumo_atual=filtros['classificacao_resumo_atual'])
    if 'esteira' in filtros:
        query &= Q(esteira=filtros['esteira'])
    
    # Aplicar filtro por segmento se fornecido
    if segmento:
        query &= Q(segmento_v3=segmento)
    
    # Aplicar filtro de busca global
    if search:
        query &= (
            Q(pedido__icontains=search) |
            Q(cliente__icontains=search) |
            Q(carteira__icontains=search) |
            Q(pendencia_macro__icontains=search) |
            Q(estimativa__icontains=search) |
            Q(segmento_v3__icontains=search)
        )
    
    # Obter total de registros
    total = BaseConsolidada.objects.filter(query).count()
    
    # Obter registros filtrados com ordenação
    order_field = column_map.get(order_column, 'pedido')
    if order_dir == 'desc':
        order_field = f'-{order_field}'
    
    registros = BaseConsolidada.objects.filter(query).order_by(order_field)
    filtered = registros.count()
    
    # Aplicar paginação
    registros = registros[start:start + length]
    
    # Preparar dados para resposta
    data = []
    for registro in registros:
        data.append({
            'pedido': registro.pedido,
            'dias_carteira_atual': registro.dias_carteira_atual or '-',
            'data_tecnica': registro.data_tecnica.strftime('%d/%m/%Y') if registro.data_tecnica else '-',
            'carteira': registro.carteira or '-',
            'pendencia_macro': registro.pendencia_macro or '-',
            'estimativa': registro.estimativa or '-',
            'segmento_v3': registro.segmento_v3 or '-'
        })
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total,
        'recordsFiltered': filtered,
        'data': data
    })

@login_required
def revisao_segmentos(request):
    tipo = request.GET.get('tipo', 'revisar')
    print(f"Tipo recebido: {tipo}")
    
    # Obter filtros do módulo
    filtros = get_filtro(tipo)
    print(f"Filtros obtidos: {filtros}")
    
    # Construir a query base
    query = Q()
    
    # Aplicar filtros baseado no tipo selecionado
    if 'pendencia_macro' in filtros:
        query &= Q(pendencia_macro__in=filtros['pendencia_macro'])
        print(f"Query após pendencia_macro: {query}")
    if 'classificacao_resumo_atual' in filtros:
        query &= Q(classificacao_resumo_atual=filtros['classificacao_resumo_atual'])
        print(f"Query após classificacao_resumo_atual: {query}")
    if 'esteira' in filtros:
        query &= Q(esteira=filtros['esteira'])
        print(f"Query após esteira: {query}")
    
    # Debug: Verificar valores únicos no banco
    pendencias_unicas = BaseConsolidada.objects.values_list('pendencia_macro', flat=True).distinct()
    classificacoes_unicas = BaseConsolidada.objects.values_list('classificacao_resumo_atual', flat=True).distinct()
    esteiras_unicas = BaseConsolidada.objects.values_list('esteira', flat=True).distinct()
    
    print(f"Pendências únicas no banco: {list(pendencias_unicas)}")
    print(f"Classificações únicas no banco: {list(classificacoes_unicas)}")
    print(f"Esteiras únicas no banco: {list(esteiras_unicas)}")
    
    # Obter total geral
    total_geral = BaseConsolidada.objects.filter(query).count()
    print(f"Total geral após filtros: {total_geral}")
    
    # Obter contagem por segmento
    segmentos = BaseConsolidada.objects.filter(query).values('segmento_v3').annotate(
        total=Count('id')
    ).order_by('segmento_v3')
    
    # Preparar resposta
    data = {
        'total_geral': total_geral,
        'segmentos': [
            {
                'segmento': item['segmento_v3'] or 'Não Informado',
                'total': item['total']
            }
            for item in segmentos
        ]
    }
    
    print(f"Resposta final: {data}")
    return JsonResponse(data)

@method_decorator(login_required, name='dispatch')
class BaseConsolidadaView(ListView):
    model = BaseConsolidada
    template_name = 'processamento/base_consolidada.html'
    context_object_name = 'registros'
    paginate_by = 50

    def get_queryset(self):
        return BaseConsolidada.objects.all().order_by('-created_at')

@login_required
def base_consolidada_data(request):
    # Obter parâmetros do DataTables
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    
    # Obter parâmetros de ordenação
    order_column = request.GET.get('order[0][column]', '0')
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    # Mapear colunas para campos do modelo
    column_map = {
        '0': 'pedido',
        '1': 'dias_carteira_atual',
        '2': 'data_tecnica',
        '3': 'carteira',
        '4': 'pendencia_macro',
        '5': 'estimativa',
        '6': 'segmento_v3'
    }
    
    # Obter parâmetros de filtro
    pedido = request.GET.get('pedido', '')
    cliente = request.GET.get('cliente', '')
    spe = request.GET.get('spe', '')
    atp = request.GET.get('atp', '')
    
    # Construir query
    query = Q()
    if pedido:
        query &= Q(pedido__icontains=pedido)
    if cliente:
        query &= Q(cliente__icontains=cliente)
    if spe:
        query &= Q(spe__icontains=spe)
    if atp:
        query &= Q(atp__icontains=atp)
    
    # Obter total de registros
    total = BaseConsolidada.objects.count()
    
    # Obter registros filtrados com ordenação
    order_field = column_map.get(order_column, 'pedido')
    if order_dir == 'desc':
        order_field = f'-{order_field}'
    
    registros = BaseConsolidada.objects.filter(query).order_by(order_field)
    filtered = registros.count()
    
    # Aplicar paginação
    registros = registros[start:start + length]
    
    # Preparar dados
    data = []
    for registro in registros:
        data.append({
            'pedido': registro.pedido,
            'dias_carteira_atual': registro.dias_carteira_atual or '-',
            'data_tecnica': registro.data_tecnica.strftime('%d/%m/%Y') if registro.data_tecnica else '-',
            'carteira': registro.carteira or '-',
            'pendencia_macro': registro.pendencia_macro or '-',
            'estimativa': registro.estimativa or '-',
            'segmento_v3': registro.segmento_v3 or '-'
        })
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total,
        'recordsFiltered': filtered,
        'data': data
    })

@login_required
def detalhe_registro(request, pedido):
    registro = get_object_or_404(BaseConsolidada, pedido=pedido)
    revisao = RegistroRevisao.objects.filter(pedido=pedido).first()
    
    context = {
        'registro': registro,
        'revisao': revisao,
    }
    
    return render(request, 'processamento/detalhe_registro.html', context)

@login_required
def revisar_registro(request, pedido):
    registro = get_object_or_404(BaseConsolidada, pedido=pedido)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        observacao = request.POST.get('observacao')
        
        revisao = RegistroRevisao.objects.create(
            pedido=pedido,
            status=status,
            observacao=observacao,
            usuario=request.user
        )
        
        return JsonResponse({'success': True, 'revisao_id': revisao.id})
    
    context = {
        'registro': registro,
    }
    
    return render(request, 'processamento/revisar_registro.html', context)

@login_required
def confirmar_revisao(request, pedido):
    revisao = get_object_or_404(RegistroRevisao, pedido=pedido)
    revisao.confirmado = True
    revisao.save()
    
    return JsonResponse({'success': True})

@login_required
def rejeitar_revisao(request, pedido):
    revisao = get_object_or_404(RegistroRevisao, pedido=pedido)
    revisao.confirmado = False
    revisao.save()
    
    return JsonResponse({'success': True})

@login_required
def get_registro_detalhes(request, pedido):
    try:
        registro = BaseConsolidada.objects.get(pedido=pedido)
        data = {
            'pedido': registro.pedido,
            'cliente': registro.cliente,
            'cidade': registro.cidade,
            'esteira': registro.esteira,
            'esteira_regionalizada': registro.esteira_regionalizada,
            'seg_resumo': registro.seg_resumo,
            'produto': registro.produto,
            'servico': registro.servico,
            'classificacao': registro.classificacao_resumo_atual,
            'cadeia_pendencias_descricao': registro.cadeia_pendencias_descricao,
            'carteira': registro.carteira,
            'dias_carteira_atual': registro.dias_carteira_atual,
            'data_tecnica': registro.data_tecnica,
            'osx': registro.osx,
            'motivo_pta_cod': registro.motivo_pta_cod,
            'num_wcd': registro.wcd,
            'tm_tec_total': registro.tm_tec_total,
            'delta_rec_liq': registro.delta_rec_liq,
            'aging_resumo': registro.aging_resumo,
            'origem_pend': registro.origem_pend,
            'segmento_v3': registro.segmento_v3,
            'sla_tecnica': registro.sla_tecnica,
            'quebra_esteira': registro.quebra_esteira,
            'projetos': registro.projetos,
            'projetos_lote': registro.projetos_lote,
            'tecnologia_report': registro.tecnologia_report,
            'draft_encontrado': registro.draft_encontrado,
            'data_criacao_draft': registro.data_criacao_draft,
            'tarefa_atual_draft': registro.tarefa_atual_draft,
            'status_rede': registro.status_rede,
            'eps_rede': registro.eps_rede,
            'data_rede': registro.data_rede,
            'tipo_entrega': registro.tipo_entrega,
            'data_sae': registro.data_sae,
            'agrupado': registro.agrupado,
            'pendencia_macro': registro.pendencia_macro,
            'estimativa': registro.estimativa
        }
        return JsonResponse({'success': True, 'registro': data})
    except BaseConsolidada.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Registro não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def get_registro_detalhes_base(request, pedido):
    try:
        registro = BaseConsolidada.objects.get(pedido=pedido)
        data = {
            'cliente': registro.cliente,
            'cidade': registro.cidade,
            'esteira': registro.esteira,
            'esteira_regionalizada': registro.esteira_regionalizada,
            'seg_resumo': registro.seg_resumo,
            'produto': registro.produto,
            'servico': registro.servico,
            'classificacao_resumo_atual': registro.classificacao_resumo_atual,
            'cadeia_pendencias_descricao': registro.cadeia_pendencias_descricao,
            'carteira': registro.carteira,
            'dias_carteira_atual': registro.dias_carteira_atual,
            'data_tecnica': registro.data_tecnica.strftime('%d/%m/%Y') if registro.data_tecnica else None,
            'osx': registro.osx,
            'motivo_pta_cod': registro.motivo_pta_cod,
            'status_rede': registro.status_rede,
            'eps_rede': registro.eps_rede,
            'data_rede': registro.data_rede.strftime('%d/%m/%Y') if registro.data_rede else None,
            'tipo_entrega': registro.tipo_entrega,
            'data_sae': registro.data_sae.strftime('%d/%m/%Y') if registro.data_sae else None,
            'agrupado': registro.agrupado,
            'pendencia_macro': registro.pendencia_macro,
            'estimativa': registro.estimativa,
            'wcd': registro.wcd,
            'num_atp': registro.num_atp,
            'draft_encontrado': registro.draft_encontrado,
            'data_criacao_draft': registro.data_criacao_draft.strftime('%d/%m/%Y') if registro.data_criacao_draft else None
        }
        return JsonResponse(data)
    except BaseConsolidada.DoesNotExist:
        return JsonResponse({'error': 'Registro não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def detalhe_revisao(request, pedido):
    registro = get_object_or_404(BaseConsolidada, pedido=pedido)
    logs_alteracao = LogAlteracaoPendenciaMacro.objects.filter(pedido=pedido).order_by('-data_alteracao')
    
    context = {
        'registro': registro,
        'pendencias_macro_opcoes': PENDENCIAS_MACRO_OPCOES,
        'logs_alteracao': logs_alteracao
    }
    
    return render(request, 'processamento/detalhe_registro.html', context)

@login_required
def atualizar_pendencia_macro(request, pedido):
    if request.method == 'POST':
        # Obter o registro a ser atualizado
        registro = get_object_or_404(BaseConsolidada, pedido=pedido)
        
        # Obter o novo valor de pendencia_macro do formulário
        nova_pendencia_macro = request.POST.get('pendencia_macro')
        
        # Registrar o log de alteração
        valor_anterior = registro.pendencia_macro
        LogAlteracaoPendenciaMacro.objects.create(
            pedido=pedido,
            valor_anterior=valor_anterior,
            valor_novo=nova_pendencia_macro,
            usuario=request.user
        )
        
        # Atualizar o registro
        registro.pendencia_macro = nova_pendencia_macro
        registro.save()
        
        # Adicionar mensagem de sucesso
        messages.success(request, f'Pendência Macro atualizada com sucesso para "{nova_pendencia_macro}".')
    
    # Redirecionar de volta para a página de detalhes
    return redirect('detalhe_revisao', pedido=pedido)

@login_required
def exportar_base_excel(request):
    try:
        start_time = time.time()
        
        # Registra o início da exportação
        log = ArquivoProcessado(
            nome="Exportação da Base Consolidada",
            tipo_arquivo="EXPORTACAO",
            status="PROCESSANDO",
            usuario=request.user
        )
        log.save()
        
        # Limpar arquivos temporários antigos
        for file in os.listdir('.'):
            if file.startswith('base_batimento_report_') and file.endswith('.xlsx'):
                try:
                    os.remove(file)
                except Exception as e:
                    print(f"Erro ao remover arquivo temporário {file}: {str(e)}")
        
        # Obtém todos os registros da base consolidada
        registros = BaseConsolidada.objects.all()
        
        # Converte para DataFrame
        df = pd.DataFrame(list(registros.values()))
        
        if len(df) == 0:
            messages.warning(request, 'Não há dados na Base Consolidada para exportar.')
            return redirect('upload_arquivos')
        
        # Remove colunas indesejadas
        colunas_remover = ['id', 'created_at', 'updated_at']
        df = df.drop(columns=[col for col in colunas_remover if col in df.columns])
        
        # Remove timezone e converte campos datetime
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    df[col] = df[col].dt.tz_localize(None)
                except Exception as e:
                    print(f"Erro ao processar coluna {col}: {str(e)}")
                    df[col] = ''
            df[col] = df[col].fillna('')
        
        # Define o caminho do arquivo modelo
        arquivo_modelo = os.path.join(settings.MEDIA_ROOT, 'modelos', 'modelo_batimento.xlsx')
        
        # Verifica se o arquivo modelo existe
        if not os.path.exists(arquivo_modelo):
            arquivo_modelo = os.path.join(BASE_DIR, 'processamento', 'modelos', 'modelo_batimento.xlsx')
            if not os.path.exists(arquivo_modelo):
                raise FileNotFoundError(f"Arquivo modelo não encontrado em {arquivo_modelo}")
        
        # Gera o nome do arquivo com timestamp
        now = datetime.now()
        formatted_date = now.strftime("%d_%m_%H_%M")
        arquivo_saida = f"base_batimento_report_{formatted_date}.xlsx"
        
        # Copia o arquivo modelo 
        import shutil
        shutil.copy2(arquivo_modelo, arquivo_saida)
        
        # Carrega o arquivo copiado
        livro = openpyxl.load_workbook(arquivo_saida)
        
        # Verifica se a planilha existe
        if 'Base_Consolidada' not in livro.sheetnames:
            raise ValueError(f"Aba 'Base_Consolidada' não encontrada no arquivo modelo.")
        
        sheet = livro['Base_Consolidada']
        
        # Limpa o conteúdo existente (exceto cabeçalhos)
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        
        # Verificar se há dados no DataFrame
        if not df.empty:
            # Garantir que temos os cabeçalhos corretos
            headers = [cell.value for cell in sheet[1]]
            
            # Se não houver cabeçalhos ou estiverem incorretos, adicioná-los
            if not headers or len(headers) != len(df.columns):
                for i, col_name in enumerate(df.columns, start=1):
                    sheet.cell(row=1, column=i, value=col_name)
            
            # Atualiza os dados
            for i, row in enumerate(df.values, start=2):
                for j, value in enumerate(row, start=1):
                    sheet.cell(row=i, column=j, value=value)
        
        # Salva o arquivo
        livro.save(arquivo_saida)
        
        # Garante que o arquivo existe
        if not os.path.exists(arquivo_saida):
            raise FileNotFoundError(f"Arquivo de saída {arquivo_saida} não foi criado corretamente.")
        
        # Lê o arquivo para download
        with open(arquivo_saida, 'rb') as f:
            file_data = f.read()
            response = HttpResponse(file_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={arquivo_saida}'
        
        # Remove o arquivo temporário depois de ter lido os dados
        try:
            os.remove(arquivo_saida)
        except Exception as e:
            print(f"Erro ao remover arquivo temporário {arquivo_saida}: {str(e)}")
        
        # Calcula o tempo de processamento
        end_time = time.time()
        tempo_processamento = end_time - start_time
        
        # Atualiza o log com o resultado
        log.status = "CONCLUIDO"
        log.mensagem = f"Exportação da base consolidada concluída com sucesso. Arquivo: {arquivo_saida}"
        log.registros_afetados = BaseConsolidada.objects.count()
        log.tempo_processamento = tempo_processamento
        log.save()
        
        return response
        
    except Exception as e:
        # Em caso de erro, registra no log
        if 'log' in locals():
            log.status = "ERRO"
            log.mensagem = f"Erro ao exportar base: {str(e)}"
            log.save()
        else:
            ArquivoProcessado.objects.create(
                nome="Exportação da Base Consolidada",
                tipo_arquivo="EXPORTACAO",
                status="ERRO",
                mensagem=f"Erro ao exportar base: {str(e)}",
                usuario=request.user
            )
        
        messages.error(request, f'Erro ao exportar base: {str(e)}')
        return redirect('upload_arquivos')

@login_required
def arquivo_detalhes(request, arquivo_id):
    """API para obter detalhes de um arquivo processado"""
    try:
        arquivo = ArquivoProcessado.objects.get(id=arquivo_id)
        
        # Obtém os valores de display para o tipo de arquivo e status
        tipo_arquivo_display = dict(ArquivoProcessado.TIPO_CHOICES).get(arquivo.tipo_arquivo, arquivo.tipo_arquivo)
        status_display = dict(ArquivoProcessado.STATUS_CHOICES).get(arquivo.status, arquivo.status)
        
        # Formata a data de processamento
        data_processamento = arquivo.data_processamento.strftime("%d/%m/%Y %H:%M:%S")
        
        return JsonResponse({
            'id': arquivo.id,
            'nome': arquivo.nome,
            'tipo_arquivo': arquivo.tipo_arquivo,
            'tipo_arquivo_display': tipo_arquivo_display,
            'status': arquivo.status,
            'status_display': status_display,
            'data_processamento': data_processamento,
            'usuario': arquivo.usuario.username,
            'mensagem': arquivo.mensagem,
            'registros_afetados': arquivo.registros_afetados,
            'tempo_processamento': arquivo.tempo_processamento,
        })
    except ArquivoProcessado.DoesNotExist:
        return JsonResponse({'error': 'Arquivo não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Função de logout personalizada que aceita métodos GET
def custom_logout(request):
    logout(request)
    messages.success(request, 'Você saiu do sistema com sucesso.')
    return redirect('login')

@login_required
def exportar_revisao_excel(request):
    try:
        start_time = time.time()
        
        # Registra o início da exportação
        log = ArquivoProcessado(
            nome="Exportação de Dados da Revisão",
            tipo_arquivo="EXPORTACAO",
            status="PROCESSANDO",
            usuario=request.user
        )
        log.save()
        
        # Limpar arquivos temporários antigos
        for file in os.listdir('.'):
            if file.startswith('revisao_report_') and file.endswith('.xlsx'):
                try:
                    os.remove(file)
                except Exception as e:
                    print(f"Erro ao remover arquivo temporário {file}: {str(e)}")
        
        # Obter parâmetros de filtro
        segmento = request.GET.get('segmento', '')
        tipo = request.GET.get('tipo', 'revisar')
        search = request.GET.get('search', '')
        
        # Obter filtros do módulo
        filtros = get_filtro(tipo)
        
        # Construir a query base
        query = Q()
        
        # Aplicar filtros baseado no tipo selecionado
        if 'pendencia_macro' in filtros:
            query &= Q(pendencia_macro__in=filtros['pendencia_macro'])
        if 'classificacao_resumo_atual' in filtros:
            query &= Q(classificacao_resumo_atual=filtros['classificacao_resumo_atual'])
        if 'esteira' in filtros:
            query &= Q(esteira=filtros['esteira'])
        
        # Aplicar filtro por segmento se fornecido
        if segmento:
            query &= Q(segmento_v3=segmento)
        
        # Aplicar filtro de busca global
        if search:
            query &= (
                Q(pedido__icontains=search) |
                Q(cliente__icontains=search) |
                Q(carteira__icontains=search) |
                Q(pendencia_macro__icontains=search) |
                Q(estimativa__icontains=search) |
                Q(segmento_v3__icontains=search)
            )
        
        # Obtém os registros filtrados
        registros = BaseConsolidada.objects.filter(query)
        
        # Converte para DataFrame
        df = pd.DataFrame(list(registros.values()))
        
        # Remove colunas indesejadas
        colunas_remover = ['id', 'created_at', 'updated_at']
        df = df.drop(columns=[col for col in colunas_remover if col in df.columns])
        
        # Remove timezone e converte campos datetime
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    df[col] = df[col].dt.tz_localize(None)
                except Exception as e:
                    print(f"Erro ao processar coluna {col}: {str(e)}")
                    df[col] = ''
            df[col] = df[col].fillna('')
        
        # Gera o nome do arquivo com timestamp
        now = datetime.now()
        formatted_date = now.strftime("%d_%m_%H_%M")
        arquivo_saida = f"revisao_report_{formatted_date}.xlsx"
        
        # Criar o arquivo Excel diretamente com pandas
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
            # Escrever o DataFrame para o Excel
            df.to_excel(writer, sheet_name='Base_Consolidada', index=False)
            
            # Ajustar larguras das colunas automaticamente
            worksheet = writer.sheets['Base_Consolidada']
            for i, col in enumerate(df.columns):
                max_len = max(
                    df[col].astype(str).map(len).max(),  # comprimento máximo do conteúdo
                    len(str(col))  # comprimento do cabeçalho
                ) + 2  # adicionar um pouco de espaço
                
                # Evitar colunas muito largas
                max_len = min(max_len, 50)
                
                # Converter índice de coluna para letra (A, B, C, ...)
                col_letter = openpyxl.utils.get_column_letter(i + 1)
                worksheet.column_dimensions[col_letter].width = max_len
        
        # Lê o arquivo para download
        with open(arquivo_saida, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={arquivo_saida}'
        
        # Remove o arquivo temporário
        try:
            os.remove(arquivo_saida)
        except Exception as e:
            print(f"Erro ao remover arquivo temporário {arquivo_saida}: {str(e)}")
        
        # Calcula o tempo de processamento
        end_time = time.time()
        tempo_processamento = end_time - start_time
        
        # Atualiza o log com o resultado
        log.status = "CONCLUIDO"
        log.mensagem = f"Exportação dos dados de revisão concluída com sucesso. Arquivo: {arquivo_saida}"
        log.registros_afetados = registros.count()
        log.tempo_processamento = tempo_processamento
        log.save()
        
        return response
        
    except Exception as e:
        # Em caso de erro, registra no log
        if 'log' in locals():
            log.status = "ERRO"
            log.mensagem = f"Erro ao exportar dados de revisão: {str(e)}"
            log.save()
        else:
            ArquivoProcessado.objects.create(
                nome="Exportação de Dados da Revisão",
                tipo_arquivo="EXPORTACAO",
                status="ERRO",
                mensagem=f"Erro ao exportar dados de revisão: {str(e)}",
                usuario=request.user
            )
        
        messages.error(request, f'Erro ao exportar dados de revisão: {str(e)}')
        return redirect('revisao')
